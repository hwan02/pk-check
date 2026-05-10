"""
Pokemon TCG API에서 카드 정보를 가져오고,
PokeAPI로 일본어 이름 변환 후 snkrdunk 시세까지 조회하여 엑셀로 저장하는 스크립트

사용법:
  python3 fetch_cards.py "name:pikachu"
  python3 fetch_cards.py "name:charizard rarity:Illustration Rare"
  python3 fetch_cards.py --no-snkrdunk "set.name:151"
"""
import re
import time
import json
import urllib.parse
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys

API_BASE = "https://api.pokemontcg.io/v2"
POKEAPI_BASE = "https://pokeapi.co/api/v2/pokemon-species"
SNKRDUNK_SEARCH = "https://snkrdunk.com/search"

# PokeAPI 이름 캐시 (영어 -> 일본어)
_ja_name_cache = {}

# 레어리티 영어 -> 일본어 약어 매핑
RARITY_MAP = {
    "Illustration Rare": "AR",
    "Special Illustration Rare": "SAR",
    "Hyper Rare": "HR",
    "Ultra Rare": "UR",
    "Secret Rare": "SR",
    "Art Rare": "AR",
    "Double Rare": "RR",
    "Rare Holo V": "V",
    "Rare Holo VSTAR": "VSTAR",
    "Rare Holo VMAX": "VMAX",
    "Rare": "R",
    "Uncommon": "U",
    "Common": "C",
    "Amazing Rare": "AR",
    "Shiny Rare": "S",
    "Trainer Gallery Rare Holo": "CHR",
    "ACE SPEC Rare": "ACE",
}


def get_japanese_name(english_name):
    """PokeAPI에서 포켓몬 일본어 이름을 가져옵니다."""
    # "ex", "V", "VSTAR" 등 접미사 제거
    base_name = re.split(r'\s+(ex|EX|V|VSTAR|VMAX|GX|Tag Team|MEGA|BREAK|δ)', english_name)[0].strip()
    # 소유격 등 제거 (예: "Rocket's Mewtwo" -> "Mewtwo")
    if "'s " in base_name:
        base_name = base_name.split("'s ")[-1]

    cache_key = base_name.lower()
    if cache_key in _ja_name_cache:
        return _ja_name_cache[cache_key]

    try:
        slug = base_name.lower().replace(" ", "-").replace(".", "").replace("'", "")
        resp = requests.get(f"{POKEAPI_BASE}/{slug}", timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            ja_names = [n for n in data["names"] if n["language"]["name"] == "ja"]
            if ja_names:
                ja_name = ja_names[0]["name"]
                _ja_name_cache[cache_key] = ja_name
                return ja_name
    except Exception:
        pass

    _ja_name_cache[cache_key] = None
    return None


def search_snkrdunk(keyword):
    """snkrdunk에서 키워드로 검색하여 가장 잘 매칭되는 상품의 가격을 반환합니다."""
    try:
        encoded = urllib.parse.quote(keyword)
        url = f"{SNKRDUNK_SEARCH}?keyword={encoded}&searchCategoryIds=6"
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept-Language": "ja,en;q=0.9",
        }
        resp = requests.get(url, headers=headers, timeout=10)
        if resp.status_code != 200:
            return None, None

        # aria-label="카드명 - ¥가격" 패턴으로 추출
        items = re.findall(r'aria-label="([^"]+?)\s*-\s*¥([\d,]+)"', resp.text)
        if not items:
            return None, None

        # 검색 키워드의 핵심 부분들
        search_parts = keyword.split()
        pokemon_name = search_parts[0]  # 예: ピカチュウex

        rarity_part = search_parts[1] if len(search_parts) > 1 else ""
        base_name = re.sub(r'(ex|EX|V|VSTAR|VMAX|GX)$', '', pokemon_name)

        def is_box_product(title):
            """싱글카드가 아닌 박스/세트 상품인지 판별합니다."""
            # 카드 넘버 [XXX 000/000] 패턴이 있으면 싱글카드
            if re.search(r'\[[A-Za-z0-9-]+ \d+/\d+\]', title):
                return False
            # 프로모 넘버 패턴도 싱글카드
            if re.search(r'\[[A-Za-z0-9-]+ \d+\]', title):
                return False
            # 이외에 박스 키워드가 있으면 박스
            for bk in ["ボックス", "BOX", "デッキ", "セット", "コレクション"]:
                if bk in title:
                    return True
            return False

        # 1순위: 정확한 이름 + 레어리티 매칭 (싱글카드)
        if rarity_part:
            for title, price_str in items:
                is_box = is_box_product(title)
                if pokemon_name in title and rarity_part in title and not is_box:
                    return int(price_str.replace(",", "")), title
        # 2순위: 정확한 이름 매칭 (싱글카드)
        for title, price_str in items:
            is_box = is_box_product(title)
            if pokemon_name in title and not is_box:
                return int(price_str.replace(",", "")), title
        # 3순위: 베이스 이름 + 레어리티 매칭 (싱글카드)
        if rarity_part and base_name != pokemon_name:
            for title, price_str in items:
                is_box = is_box_product(title)
                if base_name in title and rarity_part in title and not is_box:
                    return int(price_str.replace(",", "")), title
        # 4순위: 베이스 이름 매칭 (싱글카드)
        if base_name != pokemon_name:
            for title, price_str in items:
                is_box = is_box_product(title)
                if base_name in title and not is_box:
                    return int(price_str.replace(",", "")), title
    except Exception:
        pass
    return None, None


def fetch_cards(query=None, page_size=250, max_pages=5):
    """카드 데이터를 API에서 직접 가져옵니다."""
    all_cards = []
    for page in range(1, max_pages + 1):
        print(f"  페이지 {page} 로딩 중...")
        params = {"page": page, "pageSize": page_size}
        if query:
            params["q"] = query
        resp = requests.get(f"{API_BASE}/cards", params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        cards = data.get("data", [])
        if not cards:
            break
        all_cards.extend(cards)
        print(f"  -> {len(cards)}장 로드 (누적: {len(all_cards)}장)")
        total = data.get("totalCount", 0)
        if len(all_cards) >= total:
            break
    return all_cards


def build_snkrdunk_keyword(card, ja_name):
    """snkrdunk 검색용 키워드를 조합합니다."""
    name = card.get("name", "")
    rarity = card.get("rarity", "")

    # 접미사 추출 (ex, V, VSTAR 등)
    suffix = ""
    for s in ["VSTAR", "VMAX", "GX", "EX", "ex", " V"]:
        if s in name:
            suffix = s.strip()
            break

    # 레어리티 일본어 약어
    rarity_ja = RARITY_MAP.get(rarity, "")

    keyword = ja_name
    if suffix:
        keyword += suffix
    if rarity_ja:
        keyword += f" {rarity_ja}"

    return keyword


def get_attacks_text(card):
    attacks = card.get("attacks", []) or []
    parts = []
    for atk in attacks:
        cost = ",".join(atk.get("cost", []))
        name = atk.get("name", "")
        damage = atk.get("damage", "")
        parts.append(f"[{cost}] {name} {damage}".strip())
    return " / ".join(parts)


def get_weaknesses_text(card):
    weaknesses = card.get("weaknesses", []) or []
    return ", ".join(f"{w.get('type', '')} {w.get('value', '')}" for w in weaknesses)


def get_resistances_text(card):
    resistances = card.get("resistances", []) or []
    return ", ".join(f"{r.get('type', '')} {r.get('value', '')}" for r in resistances)


def get_tcgplayer_price(card):
    market, low, mid, high = "", "", "", ""
    tcgplayer = card.get("tcgplayer", {}) or {}
    prices = tcgplayer.get("prices", {}) or {}
    for price_type in ["normal", "holofoil", "reverseHolofoil", "1stEditionHolofoil"]:
        if price_type in prices:
            p = prices[price_type]
            market = p.get("market", "")
            low = p.get("low", "")
            mid = p.get("mid", "")
            high = p.get("high", "")
            break
    return market, low, mid, high


def create_excel(cards, snkrdunk_data, filename="pokemon_cards.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pokemon Cards"

    headers = [
        "카드ID", "이름", "이름(일본어)", "슈퍼타입", "타입", "서브타입",
        "HP", "레어리티", "세트명", "세트 시리즈", "카드번호",
        "아티스트", "공격기술", "약점", "저항력", "후퇴비용",
        "TCGPlayer($)", "TCGPlayer Low($)", "TCGPlayer Mid($)", "TCGPlayer High($)",
        "snkrdunk(¥)", "snkrdunk 상품명",
        "이미지 URL"
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, card in enumerate(cards, 2):
        card_id = card.get("id", "")
        market, low, mid, high = get_tcgplayer_price(card)
        retreat_cost = card.get("retreatCost", []) or []
        types = card.get("types", []) or []
        subtypes = card.get("subtypes", []) or []
        card_set = card.get("set", {}) or {}
        images = card.get("images", {}) or {}

        snkr = snkrdunk_data.get(card_id, {})
        snkr_price = snkr.get("price", "")
        snkr_title = snkr.get("title", "")
        ja_name = snkr.get("ja_name", "")

        row_data = [
            card_id,
            card.get("name", ""),
            ja_name,
            card.get("supertype", ""),
            ", ".join(types),
            ", ".join(subtypes),
            card.get("hp", ""),
            card.get("rarity", ""),
            card_set.get("name", ""),
            card_set.get("series", ""),
            card.get("number", ""),
            card.get("artist", ""),
            get_attacks_text(card),
            get_weaknesses_text(card),
            get_resistances_text(card),
            ", ".join(retreat_cost),
            market,
            low,
            mid,
            high,
            snkr_price,
            snkr_title,
            images.get("large", ""),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    col_widths = [12, 18, 14, 10, 12, 12, 6, 16, 20, 16, 8, 16,
                  40, 12, 12, 10, 12, 10, 10, 10, 12, 35, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(filename)
    print(f"\n  엑셀 저장 완료: {filename}")
    print(f"   총 {len(cards)}장의 카드 정보가 저장되었습니다.")


def main():
    print("=" * 50)
    print("  Pokemon TCG 카드 정보 + snkrdunk 시세 엑셀 생성기")
    print("=" * 50)

    use_snkrdunk = True
    args = sys.argv[1:]

    if "--no-snkrdunk" in args:
        use_snkrdunk = False
        args.remove("--no-snkrdunk")

    filename = "pokemon_cards.xlsx"

    if args:
        query = " ".join(args)
        print(f"\n검색 조건: {query}")
    else:
        query = "supertype:Pokémon"
        print(f"\n기본 검색: 모든 포켓몬 카드 (최대 1250장)")

    # 1) Pokemon TCG API에서 카드 데이터 가져오기
    print("\n[1/3] Pokemon TCG API에서 데이터를 가져오는 중...")
    cards = fetch_cards(query=query, page_size=250, max_pages=5)

    if not cards:
        print("카드를 찾을 수 없습니다.")
        return

    # 2) PokeAPI로 일본어 이름 변환 + snkrdunk 시세 조회
    snkrdunk_data = {}

    if use_snkrdunk:
        # 포켓몬 카드만 필터 (트레이너/에너지 제외)
        pokemon_cards = [c for c in cards if c.get("supertype") == "Pokémon"]
        print(f"\n[2/3] PokeAPI에서 일본어 이름 변환 중... ({len(pokemon_cards)}장)")

        unique_names = set()
        for card in pokemon_cards:
            name = card.get("name", "")
            base = re.split(r'\s+(ex|EX|V|VSTAR|VMAX|GX)', name)[0].strip()
            if "'s " in base:
                base = base.split("'s ")[-1]
            unique_names.add(base)

        print(f"  고유 포켓몬 이름 {len(unique_names)}개 변환 중...")
        for i, name in enumerate(unique_names):
            get_japanese_name(name)
            if (i + 1) % 20 == 0:
                print(f"  -> {i + 1}/{len(unique_names)} 변환 완료")
                time.sleep(0.5)  # PokeAPI rate limit

        translated = sum(1 for v in _ja_name_cache.values() if v)
        print(f"  -> 변환 성공: {translated}/{len(unique_names)}")

        print(f"\n[3/3] snkrdunk 시세 조회 중...")
        searched = 0
        found = 0
        for card in cards:
            card_id = card.get("id", "")
            card_name = card.get("name", "")
            ja_name = get_japanese_name(card_name)

            if ja_name:
                keyword = build_snkrdunk_keyword(card, ja_name)
                price, title = search_snkrdunk(keyword)
                snkrdunk_data[card_id] = {
                    "ja_name": ja_name,
                    "price": price or "",
                    "title": title or "",
                }
                if price:
                    found += 1
                searched += 1
                if searched % 5 == 0:
                    print(f"  -> {searched}장 조회 완료 (시세 확인: {found}건)")
                    time.sleep(1)  # snkrdunk rate limit
            else:
                snkrdunk_data[card_id] = {"ja_name": "", "price": "", "title": ""}

        print(f"  -> 총 {found}/{searched}건 시세 확인 완료")
    else:
        print("\n[2/3] snkrdunk 시세 조회 건너뜀 (--no-snkrdunk)")
        print("[3/3] 건너뜀")

    # 3) 엑셀 생성
    print(f"\n엑셀 파일 생성 중...")
    create_excel(cards, snkrdunk_data, filename)

    print(f"\n사용 예시:")
    print(f'  python3 fetch_cards.py "name:pikachu"')
    print(f'  python3 fetch_cards.py "name:charizard rarity:Illustration Rare"')
    print(f'  python3 fetch_cards.py "set.name:151"')
    print(f'  python3 fetch_cards.py --no-snkrdunk "supertype:Pokémon"')


if __name__ == "__main__":
    main()
